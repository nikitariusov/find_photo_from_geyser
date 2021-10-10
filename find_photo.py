#Проверка
import openpyxl
import xml.etree.ElementTree as ET
import re
import requests
from bs4 import BeautifulSoup
import time

list_ID = []
list_articul = []
name_of_xml_file_1 = ""
xl_file = ""
HOST = 'https://geyser.com.ua'
HEADERS = {'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                         'Chrome/90.0.4430.212 Safari/537.36', 'accept': '*/*'}


def read_exel(file_xl, xl_list):

    print(f'Количество товара: {xl_list.max_row - 1}')

    for i in range(1, xl_list.max_row):
        for cols in xl_list.iter_cols(1, 1):
            list_ID.append(cols[i].value)
            for cols in xl_list.iter_cols(13, 13):
                list_articul.append(cols[i].value)

    # создаем словарь ID товара : артикул поставщика
    data = {}
    n = 0
    for i in list_ID:
        data[i] = transform_art(list_articul[n])
        n += 1
    return data


def read_xml(xml_file):     # читаем xml и возвращаем его "дерево"
    tree = ET.parse(xml_file)
    return tree


def get_photo(tree, id_tovara):
    photos = []
    pictures = tree.findall(f"shop/offers/offer[@id='{id_tovara}']/picture")
    for photo in pictures:
        photos.append(photo.text)
    if photos == []:
        print(f'Код поставщика: {id_tovara} - фото не найдено! Выполняется поиск на сайте...')
        return get_html(id_tovara, HEADERS)
    else:
        return photos


def transform_art(ar):      # преобразуем артикул в просто цифры
    print(ar)
    if ar is None:
        return None
    else:
        nums = re.findall(r'\d*\.\d+|\d+', ar)
        nums = [int(i) for i in nums]
        return nums[0]


def get_html(art, HEADERS, params=None):
    url = f"https://geyser.com.ua/catalog/search/?q={art}"
    params = None
    r = requests.get(url, headers=HEADERS, params=params)
    html = r
    if html.status_code == 200:
        soup = BeautifulSoup(html.text, 'html.parser')
        items = soup.find_all('div', class_="catalogCard-title")
        if items == []:
            print(f'***** ВНИМАНИЕ! Товар {art} отсутвует на сайте поставщика! ***** \n')
            return ['not image']
        else:
            link = HOST + items[0].find('a').get('href')
            return search_photo_on_site(art, link, HEADERS)


def search_photo_on_site(art, link, HEADERS, params=None):
    r = requests.get(link, headers=HEADERS, params=params)
    html = r
    if html.status_code == 200:
        soup = BeautifulSoup(html.text, 'html.parser')
        item = soup.find('div', class_="gallery__photos-container")
        links_photo = item.find_all('span', class_="gallery__link j-gallery-zoom j-gallery-link")
        photo = []
        for link in links_photo:
            photo.append(HOST + link.get('data-href'))
        print(f'Товар: {art}, найдено {len(photo)} фото. \n')
        return photo


def recording_on_file(kod_list, xl_list, file_xl):
    row = 2
    for links in kod_list.values():
        links_to_save = ''
        for link in links:
            links_to_save += f'{link},'
        xl_list[f'X{row}'] = links_to_save[:-1]
        if links_to_save == 'not image,':
            xl_list[f'Y{row}'] = 0
        else:
            xl_list[f'Y{row}'] = 1
        row += 1

    file_xl.save(f'Обработан_{xl_file}')
    print(f'Запись завершена. \nФайл сохранен: "Обработан_{xl_file}"')


def main(excel_file):
    kod_and_photo = {}

    file_xl = openpyxl.load_workbook(excel_file)
    xl_list = file_xl.active

    data = read_exel(file_xl, xl_list)
    tree = read_xml(name_of_xml_file_1)

    n = 0
    for kod in data.keys():
        n += 1
        photo = get_photo(tree, data[kod])
        print(f'{n}. Поиск: код КТУ {kod} - код поставщика {data[kod]}')
        kod_and_photo[kod] = photo

    print(kod_and_photo)
    print(f'\nОбработано {len(kod_and_photo)} товаров. \nПроизводится запись в файл...')

    recording_on_file(kod_and_photo, xl_list, file_xl)


'''Приветсвие и инструкция по работе'''

print('''Привет! Данная программа предназначена для поиска фотографий на товар и сохранения ссылок в таблицу Excel.
Поиск возможен только по товарам поставщика "Гейзер"! 
Для правильной работы программы необходимо 2 файла: 
    1 - Excel файл с перечнем товара.
    2 - XML файл что выгружен из b2b Гейзера.
    
                  *** ВАЖНО ***
    Оба файла должны находиться в той же папке, 
            что и данная программа!
            
В файле Excel должны быть заполненны 2 колонки: 
    -Колонка А - ваш код товара, для удобства.
    -Колонка М - артикул поставщика, если отсутствует то будет проигнорирован программой.
После обработки будет сохранен новый файл с ссылками на фото в колонке Х, а в колонке Y будут 0 или 1.\n''')



name_of_xml_file_1 = str(input("Введите название XML файла от Гейзера с учетом расширения "
                                   "(например Гейзер.xml) : "))
xl_file = str(input("Введите название EXCEL файла с товарами для обработки"
                        " с учетом расширения (например Смесители.xlsx) : "))
start_time = time.time()

main(xl_file)
print(f"Время выполнения: {time.time() - start_time} секунд")
input('\nДля выхода нажми Enter...')
